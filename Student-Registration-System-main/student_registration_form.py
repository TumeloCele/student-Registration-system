#tkinter is GUI library, it's only framework built into the python standard library
from tkinter import *

#datetime module used for manipulating dates and times
from datetime import date

#tkinter.messagebox module is used to display message boxes in your applications
from tkinter import messagebox

#tkinter.filedialog module is used to open the file explorer window
from tkinter import filedialog

#Pillow is a fork of the Python Imaging Library (PIL)
from PIL import ImageTk, Image

#os module provides functions for interacting with the operating system
import os

#tkinter.ttk module provides access to the Tk themed widget set
from tkinter.ttk import Combobox

#openpyxl is a Python library to read/write Excel files
import openpyxl
from openpyxl import Workbook

#pathlib module offers classes representing filesystem paths with semantics appropriate for different operating systems
import pathlib


background_color = "#87CEFF"
frame_bg = "#228B22"
frame_fg = "#000000"

root = Tk()
root.title("Student Registration Form")
root.geometry("1250x700+210+80")
root.config(bg=background_color)


file = pathlib.Path("student_data.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Registration No."
    sheet["B1"] = "Name"
    sheet["C1"] = "Class"
    sheet["D1"] = "Gender"
    sheet["E1"] = "Date of Birth"
    sheet["F1"] = "Date of Registration"
    sheet["G1"] = "Religion"
    sheet["H1"] = "Occupation"
    sheet["I1"] = "Father's Name"
    sheet["J1"] = "Mother's Name"
    sheet["K1"] = "Father's Occupation"
    sheet["L1"] = "Mother's Occupation"

    file.save("student_data.xlsx")

#___Exit Window___#
def exit():
    msg = messagebox.askyesno("Exit", "Are you sure you want to exit?")
    if msg == True:
        root.destroy()

#___Show Image___#
def showImage():
    global img
    global filename
    filename = filedialog.askopenfilename(initialdir=os, title="Select a file", filetypes=(("png files", "*.png"),
                                                                                            ("jpg files", "*.jpg"), 
                                                                                            ("all files", "*.txt")))
    img = Image.open(filename)
    resized_img = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_img)
    lbl.config(image=photo2)
    lbl.image = photo2

#___Registration Number___#
def student_no():
    file = openpyxl.load_workbook("student_data.xlsx")
    sheet = file.active
    rows = sheet.max_row
    
    max_row_value = sheet.cell(row=rows, column=1).value

    try:
        student_number.set(int(max_row_value)+1)
    
    except:
        student_number.set(202301012200)

#___Clear All___#
def clear():
    Name.set("")
    DOB.set("")
    Class.set("Select Class")
    Religion.set("")
    Skill.set("")
    F_Name.set("")
    M_Name.set("")
    Father_Occupation.set("")
    Mother_Occupation.set("")
    
    student_no()

    Save_Button.config(state=NORMAL)

    img1 = PhotoImage(file="images/upload photo.png")
    lbl.config(image=img1)
    lbl.image = img1

    img = ""

#___Save Data___#
def save():
    S1 = student_number.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("error", "Select Gender")
    
    D2 = DOB.get()
    D1 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    if N1 == "" or C1 == "" or D2 == "" or Rel == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("Error", "All fields are required")
    else:
        file = openpyxl.load_workbook("student_data.xlsx")
        sheet = file.active
        sheet.cell(row=sheet.max_row+1, column=1, value = R1)
        sheet.cell(row=sheet.max_row, column=2, value = N1)
        sheet.cell(row=sheet.max_row, column=3, value = C1)
        sheet.cell(row=sheet.max_row, column=4, value = G1)
        sheet.cell(row=sheet.max_row, column=5, value = D2)
        sheet.cell(row=sheet.max_row, column=6, value = D1)
        sheet.cell(row=sheet.max_row, column=7, value = Rel)
        sheet.cell(row=sheet.max_row, column=8, value = S1)
        sheet.cell(row=sheet.max_row, column=9, value = fathername)
        sheet.cell(row=sheet.max_row, column=10, value = mothername)
        sheet.cell(row=sheet.max_row, column=11, value = F1)
        sheet.cell(row=sheet.max_row, column=12, value = M1)

        file.save("student_data.xlsx")

        try:
            img.save("student_images/"+str(R1)+".png")
        except:
            messagebox.showinfo("Info", "No Image Selected")

        messagebox.showinfo("Success", "Data Saved Successfully")
        
        clear()

        student_no()

#___Gender___#
def genderSelection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"


#top frames
Label(root, text="ayanda@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#00FF7F", fg="#fff", font=("arial", 20)).pack(side=TOP, fill=X)

#Registration Date and Number
Label(root, text="Student No:", bg=background_color, fg=frame_bg, font=("arial", 13)).place(x=30, y=150)
Label(root, text="Date:", bg=background_color, fg=frame_bg, font=("arial", 13)).place(x=500, y=150)

student_number = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=student_number, width=15, font=("arial", 10))
reg_entry.place(x=160, y=150)

student_no()

#Date of Registration
today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font=("arial", 10))
date_entry.place(x=550, y=150)

Date.set(d1)

#Student Details
obj = LabelFrame(root, text="Student's Details", width=900, height=250, bg=frame_bg, fg=frame_fg, font=("arial", 20), bd=2, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=30, y=50)
Label(obj, text="Date of Birth:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=30, y=100)
Label(obj, text="Gender:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=30, y=150)

Label(obj, text="Class:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=500, y=50)
Label(obj, text="Religion:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=500, y=100)
Label(obj, text="Skills:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font=("arial", 10))
name_entry.place(x=160, y=50)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font=("arial", 10))
dob_entry.place(x=160, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=frame_bg, fg=frame_fg, font=("arial", 10), command=genderSelection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=frame_bg, fg=frame_fg, font=("arial", 10), command=genderSelection)
R2.place(x=200, y=150)

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font=("arial", 10))
religion_entry.place(x=630, y=100)

Skill = StringVar()
skill_entry = Entry(obj, textvariable=Skill, width=20, font=("arial", 10))
skill_entry.place(x=630, y=150)

Class = Combobox(obj, values=["1st Year", "2nd Year", "3rd Year", "4th Year"], width=17, font=("Roboto", 10), state="r")
Class.place(x=630, y=50)
Class.set("Select Class")

#Parents Details
obj2 = LabelFrame(root, text="Parent's Details", width=900, height=220, bg=frame_bg, fg=frame_fg, font=("arial", 20), bd=2, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=30, y=50)
Label(obj2, text="Occupation:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=30, y=100)

F_Name = StringVar()
f_name_entry = Entry(obj2, textvariable=F_Name, width=20, font=("arial", 10))
f_name_entry.place(x=160, y=50)

Father_Occupation = StringVar()
FO_entry = Entry(obj2, textvariable=Father_Occupation, width=20, font=("arial", 10))
FO_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=500, y=50)
Label(obj2, text="Occupation:", bg=frame_bg, fg=frame_fg, font=("arial", 13)).place(x=500, y=100)

M_Name = StringVar()
M_name_entry = Entry(obj2, textvariable=M_Name, width=20, font=("arial", 10))
M_name_entry.place(x=630, y=50)

Mother_Occupation = StringVar()
MO_entry = Entry(obj2, textvariable=Mother_Occupation, width=20, font=("arial", 10))
MO_entry.place(x=630, y=100)

#Image
f = Frame(root, width=200, height=200, bg="black", bd=3, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="Images/upload photo.png")
lbl = Label(f, image=img, bg="black")
lbl.place(x=0, y=0)

#Button
Button(root, text="Upload", bg="lightblue", width=19, height=2, font=("arial", 12, "bold"), command=showImage).place(x=1000, y=370)

Save_Button = Button(root, text="Save", bg="lightgreen", width=19, height=2, font=("arial", 12, "bold"), command=save)
Save_Button.place(x=1000, y=450)

Button(root, text="Reset", bg="lightpink", width=19, height=2, font=("arial", 12, "bold"), command=clear).place(x=1000, y=530)
Button(root, text="Exit", bg="grey", width=19, height=2, font=("arial", 12, "bold"), command=exit).place(x=1000, y=610)


root.mainloop()